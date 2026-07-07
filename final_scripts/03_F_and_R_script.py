import pandas as pd
import os
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook, load_workbook
from collections import defaultdict
import math


# ─── Paths & configuration (centralized in config.py) ─────────────────────────
from config import (
    OUTPUT_DIR,
    PR_DIR,
    OVERVIEW_FILE as overview_file,
    F_R_OUTPUT_FILE as f_r_output_file,
    CURRENT_OPTION_PERIOD,
)


def _clean(val):
    """Normalize value to a clean string; blanks / NaN / 'nan' / 'none' become ''."""
    if val is None:
        return ""
    if isinstance(val, float) and math.isnan(val):
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none", "<na>"):
        return ""
    return s


def get_fr_pr_numbers(overview_path: Path):
    """
    Reads the overview file and returns a list of tuples for rows where 'F&R Needed' == 'Yes'.
    Each tuple contains:
        (PR#, Version, OpDiv, SF30 Description, 12M+ CLIN)
    """
    # Read Excel and strip column names
    df = pd.read_excel(overview_path, dtype=str)  # read all as string to preserve formatting
    df.columns = df.columns.str.strip()

    # Filter rows where F&R Needed == 'Yes'
    fr_needed = df[df["F&R Needed"].astype(str).str.strip().str.lower() == "yes"]

    # Strip whitespace and get required columns
    pr_version_list = list(
        zip(
            fr_needed["PR#"].map(_clean),
            fr_needed["Version"].map(_clean),
            fr_needed["OpDiv"].map(_clean) if "OpDiv" in fr_needed.columns else [""] * len(fr_needed),
            fr_needed["SF30 Description"].map(_clean) if "SF30 Description" in fr_needed.columns else [""] * len(fr_needed),
            fr_needed["12M+ CLIN"].map(_clean) if "12M+ CLIN" in fr_needed.columns else [""] * len(fr_needed),
        )
    )

    print(f"Found {len(pr_version_list)} PRs needing F&R.")
    return pr_version_list


def extract_comps_data(comps_df):
    """
    Extract all relevant F&R fields from the Comps worksheet in order:
    - Verizon Response
    - Source Info
    - Case Number
    - Verizon Case Description
    - Pricing Element
    - Comp Rate (as float)
    Returns a list of dicts, one per row.
    """
    records = []
    num_rows = len(comps_df)

    for i in range(num_rows):
        verizon_response = comps_df.get("Verizon's Response/HHS Comment", [""] * num_rows)[i] if "Verizon's Response/HHS Comment" in comps_df else ""
        source_info = comps_df.get("Source or Networx Information", [""] * num_rows)[i] if "Source or Networx Information" in comps_df else ""
        case_number = comps_df.get("Case Number", [""] * num_rows)[i] if "Case Number" in comps_df else ""
        verizon_case_desc = comps_df.get("Verizon Case Description", [""] * num_rows)[i] if "Verizon Case Description" in comps_df else ""
        pricing_element = comps_df.get("SRE Pricing Element", [""] * num_rows)[i] if "SRE Pricing Element" in comps_df else ""
        comp_rate_raw = comps_df.get("Comp Rate", [""] * num_rows)[i] if "Comp Rate" in comps_df else ""
        
        # Convert Comp Rate to float
        comp_rate = None
        if comp_rate_raw and comp_rate_raw != "":
            try:
                comp_rate = float(str(comp_rate_raw).replace('$', '').replace(',', ''))
            except (ValueError, TypeError):
                comp_rate = None

        record = {
            "Verizon's Response/HHS Comment": _clean(verizon_response),
            "Source or Networx Information": _clean(source_info),
            "Case Number": _clean(case_number),
            "Verizon Case Description": _clean(verizon_case_desc),
            "Pricing Element": _clean(pricing_element),
            "Comp Rate": comp_rate,  # Now a float or None
        }

        records.append(record)

    return records


def format_currency(value):
    """Format a numeric value as currency with 6 decimal places."""
    if value is None:
        return ""
    try:
        return f"${float(value):,.6f}"
    except (ValueError, TypeError):
        return ""


def get_comps_sheet(pr_file: Path):
    """Load the 'comps' worksheet (case-insensitive) and return as DataFrame."""
    try:
        xls = pd.ExcelFile(pr_file)
        comps_sheet = next((s for s in xls.sheet_names if "comp" in s.lower()), None)
        if comps_sheet:
            df = pd.read_excel(xls, sheet_name=comps_sheet)
            df.columns = df.columns.str.strip()
            return df
        else:
            print(f"  No 'comps' sheet found in {pr_file.name}")
            return pd.DataFrame()
    except Exception as e:
        print(f"  Error reading comps sheet from {pr_file.name}: {e}")
        return pd.DataFrame()



def determine_12m_clin(verizon_case_desc):
    """
    Determine 12M+ CLIN based on Verizon Case Description.
    Returns 'Yes' if '.ANN.' is found in the description, otherwise 'No'.
    """
    if verizon_case_desc is None:
        return "No"
    
    desc_str = str(verizon_case_desc).strip()
    return "Yes" if ".ANN." in desc_str.upper() else "No"



def get_j1_rate(pr_file: Path, case_number: str, pricing_element: str):
    """
    Extract J1 rate from the J1 worksheet for a given case number and pricing element.
    Returns the HHS Price as a float for the row where TO Period == "OPT PD {CURRENT_OPTION_PERIOD}".
    """
    try:
        wb = load_workbook(pr_file)
        j1_sheet = next((s for s in wb.sheetnames if "j1" in s.lower() or "j.1" in s.lower()), None)
        
        if not j1_sheet:
            print(f"  No 'J1' sheet found in {pr_file.name}")
            wb.close()
            return None
        
        ws = wb[j1_sheet]
        
        # Get headers from first row
        headers = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = idx
        
        # Check if required columns exist
        if "Case Number" not in headers or "SRE Pricing Element" not in headers or "TO Period" not in headers or "HHS Price" not in headers:
            print(f"  Missing required columns in J1 sheet of {pr_file.name}")
            wb.close()
            return None
        
        case_col = headers["Case Number"]
        pricing_col = headers["SRE Pricing Element"]
        period_col = headers["TO Period"]
        price_col = headers["HHS Price"]
        
        # Search for matching row
        for row in ws.iter_rows(min_row=2):
            row_case = str(row[case_col - 1].value).strip() if row[case_col - 1].value else ""
            raw_pricing = str(row[pricing_col - 1].value).strip() if row[pricing_col - 1].value else ""
            try:
                row_pricing = str(int(float(raw_pricing))).zfill(2) if raw_pricing else ""
            except (ValueError, TypeError):
                row_pricing = raw_pricing.zfill(2) if raw_pricing else ""
            row_period = str(row[period_col - 1].value).strip() if row[period_col - 1].value else ""
            
            # Match case number, pricing element, and TO Period
            if row_case == case_number and row_pricing == pricing_element and row_period == f"OPT PD {CURRENT_OPTION_PERIOD}":
                hhs_price_cell = row[price_col - 1]
                
                if hhs_price_cell.value is None:
                    wb.close()
                    return None
                
                # Return as float
                if isinstance(hhs_price_cell.value, (int, float)):
                    wb.close()
                    return float(hhs_price_cell.value)
                else:
                    # Try to parse string value
                    try:
                        value_str = str(hhs_price_cell.value).replace('$', '').replace(',', '')
                        wb.close()
                        return float(value_str)
                    except (ValueError, TypeError):
                        wb.close()
                        return None
        
        wb.close()
        return None
        
    except Exception as e:
        print(f"  Error reading J1 sheet from {pr_file.name}: {e}")
        return None


def build_FR():
    """
    Builds the F&R Overview Excel file replicating the official example formatting,
    converts Pricing Element to numeric when possible, and merges duplicate Case Numbers.
    Also creates separate PR tabs before merging.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Step 1: Get PRs needing F&R
    fr_pr_list = get_fr_pr_numbers(overview_file)
    fr_overview_records = []
    pr_to_file = {}

    # Step 2: Gather data
    for pr, version, opdiv, sf30_desc, cl12m in fr_pr_list:
        pr_identifier = pr if "PR" in pr.upper() else f"PR{pr}"
        print(f"Processing {pr_identifier}...")

        pr_files = [
            f for f in PR_DIR.glob(f"{pr_identifier}*.xlsx")
            if not f.name.startswith("~$")
        ]
        if not pr_files:
            print(f"  No PR files found for {pr_identifier} in {PR_DIR}")
            continue

        for pr_file in pr_files:
            comps_df = get_comps_sheet(pr_file)
            if comps_df.empty:
                continue

            if pr not in pr_to_file:
                pr_to_file[pr] = pr_file

            comps_records = extract_comps_data(comps_df)

            for rec in comps_records:
                case_num = rec.get("Case Number", "")
                case_num_str = str(case_num).strip() if case_num and not (isinstance(case_num, float) and math.isnan(case_num)) else ""

                price_elem_raw = rec.get("Pricing Element", "")
                price_elem_str = ""
                if price_elem_raw:
                    try:
                        price_elem_str = str(int(float(price_elem_raw))).zfill(2)
                    except (ValueError, TypeError):
                        price_elem_str = str(price_elem_raw).zfill(2)

                j1_rate_float = get_j1_rate(pr_file, case_num_str, price_elem_str)
                comp_rate_float = rec.get("Comp Rate")

                delta_float = None
                if j1_rate_float is not None and comp_rate_float is not None:
                    delta_float = j1_rate_float - comp_rate_float

                j1_rate = format_currency(j1_rate_float)
                comp_rate = format_currency(comp_rate_float)
                delta = format_currency(delta_float)

                fr_overview_records.append({
                    "Type of F&R": "",
                    "Verizon's Response/HHS Comment": rec.get("Verizon's Response/HHS Comment", ""),
                    "J.1 Rate": j1_rate,
                    "Comp Rate": comp_rate,
                    "Delta": delta,
                    "Source or Networx Information": rec.get("Source or Networx Information", ""),
                    "Case Number": rec.get("Case Number", ""),
                    "Verizon Case Description": rec.get("Verizon Case Description", ""),
                    "Pricing Element": price_elem_str,
                    "PR#": pr,
                    "Version": str(version).zfill(2),
                    "OpDiv": opdiv,
                    "SF30 Description": sf30_desc,
                    "12M+ CLIN": determine_12m_clin(rec.get("Verizon Case Description", ""))
                })

    if not fr_overview_records:
        print("No F&R records found.")
        return None

    # Create workbook early (before merging) to build PR tabs
    wb = Workbook()

    # --- Shared Styles ---
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    black_font = Font(bold=True, color="000000")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    # --- Define columns ---
    columns = [
        "Type of F&R",
        "Verizon's Response/HHS Comment",
        "J.1 Rate",
        "Comp Rate",
        "Delta",
        "Source or Networx Information",
        "Case Number",
        "Verizon Case Description",
        "Pricing Element",
        "PR#",
        "Version",
        "OpDiv",
        "SF30 Description",
        "12M+ CLIN"
    ]

    # Group F&R records by PR for later use in xlwings PR tab creation
    pr_groups = defaultdict(list)
    for rec in fr_overview_records:
        pr_groups[rec["PR#"]].append(rec)

    # Remove default "Sheet"
    default_ws = wb.active
    wb.remove(default_ws)

    # Merge duplicates and create Overview tab
    merged_records = []
    merge_key_columns = ("Case Number", "Pricing Element")
    concat_columns = {
        "PR#", "Version", "OpDiv", "SF30 Description",
        "12M+ CLIN", "Verizon's Response/HHS Comment", "Source or Networx Information"
    }

    temp_dict = defaultdict(list)
    for record in fr_overview_records:
        key = (record["Case Number"], record["Pricing Element"])
        temp_dict[key].append(record)

    for (case_number, pricing_element), records in temp_dict.items():
        merged_record = {}
        for col in records[0].keys():
            if col in merge_key_columns:
                merged_record[col] = records[0][col]
            elif col in concat_columns:
                values = [str(r[col]).strip() for r in records if str(r[col]).strip()]
                unique_values = list(dict.fromkeys(values))
                merged_record[col] = "/".join(unique_values) if unique_values else ""
            else:
                values = [str(r[col]).strip() for r in records if str(r[col]).strip()]
                merged_record[col] = max(values, key=len) if values else ""
        merged_records.append(merged_record)

    fr_overview_records = merged_records

    # Create Overview tab
    ws = wb.create_sheet(title="Overview")
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if col_idx <= 6:
            cell.fill = green_fill
            cell.font = black_font
        elif 7 <= col_idx <= 9:
            cell.fill = black_fill
            cell.font = white_font
        elif 10 <= col_idx <= 11:
            cell.fill = gray_fill
            cell.font = black_font
        elif 12 <= col_idx <= 13:
            cell.fill = yellow_fill
            cell.font = black_font
        elif col_idx == 14:
            cell.fill = light_green_fill
            cell.font = black_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, record in enumerate(fr_overview_records, start=2):
        for col_idx, header in enumerate(columns, start=1):
            value = record.get(header, "")
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            if 10 <= col_idx <= 11:
                c.fill = gray_fill
            elif 12 <= col_idx <= 13:
                c.fill = yellow_fill
            elif col_idx == 14:
                c.fill = light_green_fill
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, min(max_len + 3, 45))
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(columns)).coordinate}"
    ws.freeze_panes = "A2"

    wb.save(f_r_output_file)
    print(f"\nOverview tab created. Now adding PR tabs with images via xlwings...")

    # Step 3: Use xlwings to create PR tabs by copying COMP sheets from PR files
    _create_pr_tabs_xlwings(pr_groups, pr_to_file, columns[:9], f_r_output_file)

    print(f"\nF&R Overview document created with PR tabs and merged duplicates: {f_r_output_file}")
    return f_r_output_file


def _create_pr_tabs_xlwings(pr_groups, pr_to_file, fr_headers, output_file):
    """
    For each PR needing F&R:
    1. Copy the COMP sheet from the PR file into f_r_output as a new tab
    2. Delete the original Comps header + data rows
    3. Insert F&R header and data rows at the top
    Images in the COMP sheet are preserved automatically.
    """
    import xlwings as xw

    currency_cols = {"J.1 Rate", "Comp Rate", "Delta"}
    currency_fmt = '$#,##0.000000'

    def _parse_currency(val):
        """Strip '$' and commas to recover the raw float."""
        if not val or not isinstance(val, str):
            return val
        stripped = val.replace("$", "").replace(",", "").strip()
        if not stripped:
            return ""
        try:
            return round(float(stripped), 6)
        except (ValueError, TypeError):
            return val

    app = xw.App(visible=False, add_book=False)
    try:
        wb_output = app.books.open(str(output_file))

        for pr, records in pr_groups.items():
            pr_file = pr_to_file.get(pr)
            if not pr_file:
                print(f"  No PR file found for {pr}, skipping PR tab")
                continue

            pr_version = records[0]["Version"]
            tab_name = f"{pr} v{pr_version}"
            print(f"  Creating tab: {tab_name}")

            # Open PR file and find Comps sheet
            wb_pr = app.books.open(str(pr_file), read_only=True)
            comps_sheet = None
            for sht in wb_pr.sheets:
                if "comp" in sht.name.lower():
                    comps_sheet = sht
                    break

            if comps_sheet is None:
                print(f"    No COMP sheet found in {pr_file.name}")
                wb_pr.close()
                continue

            # Copy the COMP sheet to the output workbook
            comps_sheet.copy(after=wb_output.sheets[-1])
            wb_pr.close()

            # Rename the copied sheet
            new_sheet = wb_output.sheets[-1]
            new_sheet.name = tab_name

            # Find the end of the original Comps data (first fully empty row)
            used = new_sheet.used_range
            comps_data_end = 0
            for row_num in range(1, used.rows.count + 1):
                row_vals = new_sheet.range((row_num, 1), (row_num, used.columns.count)).value
                if row_vals is None:
                    break
                if isinstance(row_vals, list):
                    if all(v is None for v in row_vals):
                        break
                comps_data_end = row_num

            # Delete the original Comps data rows (header + data)
            if comps_data_end > 0:
                new_sheet.range(f"1:{comps_data_end}").delete()

            # Insert blank rows at the top for F&R header + data
            num_fr_rows = 1 + len(records)  # 1 header + N data rows
            new_sheet.range(f"1:{num_fr_rows}").insert(shift='down')

            # Write F&R header row
            for col_idx, header in enumerate(fr_headers):
                cell = new_sheet.range((1, col_idx + 1))
                cell.value = header
                cell.font.bold = True
                if col_idx < 6:
                    cell.color = (146, 208, 80)  # green
                else:
                    cell.color = (0, 0, 0)  # black
                    cell.font.color = (255, 255, 255)  # white text

            # Write F&R data rows
            for row_idx, record in enumerate(records, start=2):
                for col_idx, header in enumerate(fr_headers):
                    value = record.get(header, "")
                    cell = new_sheet.range((row_idx, col_idx + 1))
                    if header in currency_cols:
                        cell.value = _parse_currency(value)
                        if cell.value != "" and cell.value is not None:
                            cell.number_format = currency_fmt
                    else:
                        cell.value = value

            # Apply borders to F&R header + data range
            last_row = 1 + len(records)
            last_col = len(fr_headers)
            fr_range = new_sheet.range((1, 1), (last_row, last_col))
            for border_id in range(7, 13):
                fr_range.api.Borders(border_id).LineStyle = 1  # xlContinuous
                fr_range.api.Borders(border_id).Weight = 2     # xlThin

            # Auto-fit columns
            new_sheet.autofit('c')

        wb_output.save()
        wb_output.close()
    finally:
        app.quit()


if __name__ == "__main__":
    build_FR()
